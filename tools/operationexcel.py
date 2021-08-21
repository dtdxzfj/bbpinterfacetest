#!/anaconda3/envs/FEALPy/bin python3.8
# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @File: operationexcel.py
# @Author: ZFJ
# @Time: 4月 11, 2021
# ---

from openpyxl import load_workbook
from tools.pathconfig import PathConfig

"""
xlsx文件简介
xlsx是一个文档压缩格式，可以使用压缩软件打开查看组成
xlsx文件-工作表-sheet页-表格
load_workbook，载入xlsx文件，返回字典类型的sheet页

本次操作的文件只有这一个文件和一个sheet页，所以都写成默认值
row 行
column 列
max_row 最大行
以上需要手动写，没有提示
"""


class OperationExcel:

    def __init__(self, xlsxfile=PathConfig.excelpath):
        self.file = xlsxfile
        self.wb = load_workbook(self.file)  # 由于xlsx是一种excel压缩文件，打开之后需要获取工作表

    def readexcel(self, row, column, sheetname='data'):
        sheet = self.wb[sheetname]
        return sheet.cell(row, column).value

    def writeexcel(self, row, column, data, sheetname='data'):
        sheet = self.wb[sheetname]
        sheet.cell(row, column,data)
        # sheet.cell(row,column).value = data
        self.wb.save(self.file)

    # @classmethod
    def readsheet(self,sheetname  = 'data'):
        sheet = self.wb[sheetname]
        maxrow = sheet.max_row
